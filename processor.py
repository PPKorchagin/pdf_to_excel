import os
import re
from typing import Optional, List, Tuple, Callable

import pandas as pd
from PyPDF2 import PdfReader
import tabula

DOCNO_RE_1 = re.compile(r"Номер\s+документа.*?(\d{4,9})", re.IGNORECASE | re.DOTALL)
DOCNO_RE_2 = re.compile(r"(\d{1,9})\s+\d{2}\.\d{2}\.\d{4}")


def extract_doc_number(pdf_path: str) -> Optional[str]:
    reader = PdfReader(pdf_path)
    text = "\n".join((page.extract_text() or "") for page in reader.pages)

    m = DOCNO_RE_1.search(text)
    if m:
        return m.group(1)

    m = DOCNO_RE_2.search(text)
    return m.group(1) if m else None


def _choose_and_clean_table(tables: List[pd.DataFrame]) -> Optional[pd.DataFrame]:
    if not tables:
        return None

    slice_rules = {0: (3, -2), 2: (1, -2)}
    default_slice = (3, -2)

    for i, tbl in enumerate(tables):
        if tbl is None or tbl.empty:
            continue
        if len(tbl) <= 5 or len(tbl.columns) <= 2:
            continue

        start, end = slice_rules.get(i, default_slice)
        df = tbl.iloc[start:end].copy()
        if df.empty or len(df) < 2:
            continue

        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
        return df

    return None


def extract_table(pdf_path: str) -> Optional[pd.DataFrame]:
    tables = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True)
    return _choose_and_clean_table(tables)


def clean_text_series(s: pd.Series) -> pd.Series:
    s = s.astype("string").fillna("").str.replace(r"\s+", " ", regex=True).str.strip()
    s = s.replace({"": pd.NA})
    return s


def clean_code_series(s: pd.Series) -> pd.Series:
    s = s.astype("string").fillna("").str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.replace({"": pd.NA})
    return s


def _to_number_series(s: pd.Series) -> pd.Series:
    """
    Преобразует строковые значения вида:
      "1 234,56" -> 1234.56
      "1234" -> 1234
      ""/текст -> NaN
    """
    x = s.astype("string").fillna("")
    x = x.str.replace("\u00A0", " ", regex=False)  # NBSP
    x = x.str.replace(" ", "", regex=False)
    x = x.str.replace(",", ".", regex=False)
    return pd.to_numeric(x, errors="coerce")


def build_doc_df(pdf_path: str) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    Возвращает df формата:
      код | наименование | <doc_number>

    ВАЖНО: строки с одинаковым "код" внутри одного PDF суммируются.
    """
    doc_number = extract_doc_number(pdf_path)
    if not doc_number:
        return None, None

    df = extract_table(pdf_path)
    if df is None:
        return None, doc_number

    required = ["2", "3", "10"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{pdf_path}: нет колонок {missing}. Доступны: {list(df.columns)}")

    df = df.loc[:, required].copy()
    df.columns = ["наименование", "код", doc_number]

    df["наименование"] = clean_text_series(df["наименование"])
    df["код"] = clean_code_series(df["код"])
    df[doc_number] = clean_text_series(df[doc_number])

    df = df.dropna(subset=["код"])

    # значения документа -> число, далее суммируем по одинаковому коду
    df[doc_number] = _to_number_series(df[doc_number])

    df = (
        df.groupby("код", as_index=False)
        .agg({
            "наименование": lambda x: x.dropna().iloc[0] if len(x.dropna()) else pd.NA,
            doc_number: "sum",
        })
    )

    return df, doc_number


def add_totals(final_df: pd.DataFrame) -> pd.DataFrame:
    """
    Добавляет:
      - колонку ИТОГО_СТРОКА (сумма по строке по всем документным колонкам)
      - строку ИТОГО_СТОЛБЕЦ (суммы по столбцам)
    Перед добавлением итогов сортирует по "код" по возрастанию.
    """
    if final_df is None or final_df.empty:
        return final_df

    # сортировка по коду (пытаемся как число, иначе как строка)
    df = final_df.copy()
    code_num = pd.to_numeric(df["код"].astype("string"), errors="coerce")
    df = df.assign(_code_num=code_num)
    df = df.sort_values(by=["_code_num", "код"], ascending=True, na_position="last").drop(columns=["_code_num"])

    doc_cols = [c for c in df.columns if c not in ("наименование", "код")]
    num = pd.DataFrame({c: pd.to_numeric(df[c], errors="coerce") for c in doc_cols})

    df["ИТОГО_СТРОКА"] = num.sum(axis=1, skipna=True)

    totals_row = {"наименование": "ИТОГО_СТОЛБЕЦ", "код": pd.NA}
    for c in doc_cols:
        totals_row[c] = num[c].sum(skipna=True)
    totals_row["ИТОГО_СТРОКА"] = df["ИТОГО_СТРОКА"].sum(skipna=True)

    return pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)



def process_pdfs(file_paths: List[str], log: Optional[Callable[[str], None]] = None) -> pd.DataFrame:
    """
    Обрабатывает несколько PDF:
      - для каждого PDF строит таблицу по "коду"
      - объединяет по коду (outer)
      - добавляет итоги
    """
    def _log(msg: str):
        if log:
            log(msg)

    _log(f"Файлов получено: {len(file_paths)}")

    doc_dfs: List[pd.DataFrame] = []

    for path in file_paths:
        _log(f"Обработка: {os.path.basename(path)}")

        dfi, docno = build_doc_df(path)
        if docno is None:
            _log("  WARN: не найден номер документа, файл пропущен")
            continue
        if dfi is None:
            _log(f"  WARN: таблица не извлечена (doc={docno}), файл пропущен")
            continue

        _log(f"  OK: doc={docno}, уникальных кодов={len(dfi)}")
        doc_dfs.append(dfi)

    if not doc_dfs:
        final_df = pd.DataFrame(columns=["наименование", "код"])
        final_df = add_totals(final_df)
        _log("Нет данных для объединения.")
        return final_df

    # 1) Собираем "справочник" код -> наименование (первое непустое)
    names = pd.concat([d[["код", "наименование"]] for d in doc_dfs], ignore_index=True)
    names = names.dropna(subset=["код"])
    names = (
        names.groupby("код", as_index=False)["наименование"]
        .agg(lambda x: x.dropna().iloc[0] if len(x.dropna()) else pd.NA)
    )

    # 2) Объединяем документные колонки по коду
    merged = names[["код", "наименование"]].copy()
    for dfi in doc_dfs:
        doc_cols = [c for c in dfi.columns if c not in ("код", "наименование")]
        if not doc_cols:
            continue
        doc_col = doc_cols[0]
        merged = merged.merge(dfi[["код", doc_col]], on="код", how="outer")

    # порядок колонок
    doc_cols_all = [c for c in merged.columns if c not in ("код", "наименование")]
    final_df = merged[["наименование", "код"] + doc_cols_all]

    final_df = add_totals(final_df)
    _log(f"Готово. Итоговых строк (вкл. итоги): {len(final_df)}")
    return final_df


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    import io
    from openpyxl.utils import get_column_letter

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="result")

        ws = writer.book["result"]

        # авто-ширина по максимальной длине значения в колонке
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name)) if col_name is not None else 0

            for row_idx in range(2, len(df) + 2):  # данные с 2-й строки
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))

            # немного запаса, и ограничим адекватными рамками
            width = min(max_len + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    bio.seek(0)
    return bio.read()

